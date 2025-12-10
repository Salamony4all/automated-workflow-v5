"""
Excel File Processor
Handles reading and extracting data from Excel files (.xls, .xlsx)
Includes image extraction from cells (PNG, JPEG, GIF, BMP, WMF, EMF)
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import os
import json
import logging
import base64
import io
import tempfile
import subprocess
import platform
from io import BytesIO
import shutil
import zipfile
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)


def extract_wmf_images_from_excel_zip(excel_path, output_dir, sheet_name='sheet1'):
    """
    Extract WMF/EMF images directly from Excel ZIP structure with row mapping
    This bypasses openpyxl's filtering of WMF images
    
    Args:
        excel_path: Path to the Excel file
        output_dir: Directory to save extracted WMF images
        sheet_name: Sheet name (default 'sheet1')
        
    Returns:
        dict: Mapping of row numbers to WMF image paths
    """
    wmf_row_mapping = {}
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            # Parse the drawing XML to get image positions
            drawing_xml_path = f'xl/drawings/drawing1.xml'
            image_positions = {}
            
            try:
                if drawing_xml_path in zip_ref.namelist():
                    drawing_xml = zip_ref.read(drawing_xml_path).decode('utf-8')
                    root = ET.fromstring(drawing_xml)
                    
                    # Define namespaces
                    ns = {
                        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                    }
                    
                    # Find all two-cell anchors (images positioned in cells)
                    for anchor in root.findall('.//xdr:twoCellAnchor', ns):
                        # Get the starting row
                        from_elem = anchor.find('xdr:from', ns)
                        if from_elem is not None:
                            row_elem = from_elem.find('xdr:row', ns)
                            if row_elem is not None:
                                row_num = int(row_elem.text) + 1  # Convert to 1-indexed
                                
                                # Get the image reference
                                pic = anchor.find('.//xdr:pic', ns)
                                if pic is not None:
                                    blipFill = pic.find('.//xdr:blipFill', ns)
                                    if blipFill is not None:
                                        blip = blipFill.find('.//a:blip', ns)
                                        if blip is not None:
                                            embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                            if embed_id:
                                                image_positions[embed_id] = row_num
                    
                    # Parse drawing rels to map embed IDs to image files
                    drawing_rels_path = 'xl/drawings/_rels/drawing1.xml.rels'
                    if drawing_rels_path in zip_ref.namelist():
                        rels_xml = zip_ref.read(drawing_rels_path).decode('utf-8')
                        rels_root = ET.fromstring(rels_xml)
                        
                        rel_ns = {'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                        
                        for relationship in rels_root.findall('.//rel:Relationship', rel_ns):
                            rel_id = relationship.get('Id')
                            target = relationship.get('Target')
                            
                            if rel_id in image_positions and target:
                                # Extract image file name from target
                                image_file = target.split('/')[-1]
                                row_num = image_positions[rel_id]
                                
                                # Now extract and convert WMF/EMF files
                                media_path = f'xl/media/{image_file}'
                                if media_path in zip_ref.namelist() and \
                                   (image_file.lower().endswith('.wmf') or image_file.lower().endswith('.emf')):
                                    
                                    wmf_data = zip_ref.read(media_path)
                                    
                                    # Generate output filename with row info
                                    output_filename = f"{sheet_name.replace(' ', '_')}_row{row_num}_wmf_{image_file}.png"
                                    output_path = os.path.join(output_dir, output_filename)
                                    
                                    if convert_wmf_emf_to_png(wmf_data, output_path):
                                        # Store relative path for the mapping
                                        rel_path = f"imgs/{output_filename}"
                                        
                                        if row_num not in wmf_row_mapping:
                                            wmf_row_mapping[row_num] = []
                                        wmf_row_mapping[row_num].append(rel_path)
                                        
                                        logger.info(f"✓ Extracted WMF/EMF at row {row_num}: {image_file}")
                                    else:
                                        logger.warning(f"✗ Failed to convert WMF/EMF: {image_file}")
                        
            except Exception as xml_error:
                logger.debug(f"Could not parse drawing XML: {xml_error}")
                # Fall back to extracting without row mapping
                for file_info in zip_ref.filelist:
                    if file_info.filename.startswith('xl/media/') and \
                       (file_info.filename.lower().endswith('.wmf') or 
                        file_info.filename.lower().endswith('.emf')):
                        
                        wmf_data = zip_ref.read(file_info.filename)
                        base_name = os.path.basename(file_info.filename)
                        output_path = os.path.join(output_dir, f"wmf_{base_name}.png")
                        
                        if convert_wmf_emf_to_png(wmf_data, output_path):
                            logger.info(f"✓ Extracted WMF/EMF (no row mapping): {file_info.filename}")
                        
    except Exception as e:
        logger.error(f"Error extracting WMF images from ZIP: {e}")
    
    return wmf_row_mapping


def convert_wmf_emf_to_png(image_data, output_path):
    """
    Convert WMF/EMF image data to PNG format
    
    Args:
        image_data: Raw WMF/EMF image bytes
        output_path: Path where PNG should be saved
        
    Returns:
        bool: True if conversion successful, False otherwise
    """
    try:
        # Try PIL first (works for some WMF files on Windows)
        try:
            img = Image.open(BytesIO(image_data))
            img.save(output_path, 'PNG')
            logger.info(f"✓ Converted WMF/EMF to PNG using PIL")
            return True
        except Exception as pil_error:
            logger.debug(f"PIL conversion failed: {pil_error}")
        
        # Try ImageMagick (if available)
        if platform.system() in ['Linux', 'Darwin']:  # Unix-like systems
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.wmf') as tmp:
                    tmp.write(image_data)
                    tmp_path = tmp.name
                
                result = subprocess.run(
                    ['convert', tmp_path, output_path],
                    capture_output=True,
                    timeout=10
                )
                
                os.unlink(tmp_path)
                
                if result.returncode == 0 and os.path.exists(output_path):
                    logger.info(f"✓ Converted WMF/EMF to PNG using ImageMagick")
                    return True
            except Exception as im_error:
                logger.debug(f"ImageMagick conversion failed: {im_error}")
        
        # Try LibreOffice conversion (if available)
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.wmf') as tmp:
                tmp.write(image_data)
                tmp_path = tmp.name
            
            # LibreOffice can convert WMF to PNG
            result = subprocess.run(
                ['soffice', '--headless', '--convert-to', 'png', '--outdir', 
                 os.path.dirname(output_path), tmp_path],
                capture_output=True,
                timeout=15
            )
            
            if result.returncode == 0:
                # LibreOffice creates PNG with same base name
                expected_png = os.path.join(
                    os.path.dirname(output_path),
                    os.path.splitext(os.path.basename(tmp_path))[0] + '.png'
                )
                if os.path.exists(expected_png):
                    if expected_png != output_path:
                        shutil.move(expected_png, output_path)
                    os.unlink(tmp_path)
                    logger.info(f"✓ Converted WMF/EMF to PNG using LibreOffice")
                    return True
            
            os.unlink(tmp_path)
        except Exception as lo_error:
            logger.debug(f"LibreOffice conversion failed: {lo_error}")
        
        return False
        
    except Exception as e:
        logger.error(f"All WMF/EMF conversion methods failed: {e}")
        return False


class ExcelProcessor:
    """Process Excel files and extract table data with images"""
    
    def __init__(self, filepath):
        """
        Initialize Excel processor
        
        Args:
            filepath: Path to Excel file
        """
        self.original_filepath = filepath
        self.original_extension = os.path.splitext(filepath)[1].lower()
        
        # Convert .xls to .xlsx for uniform processing
        if self.original_extension == '.xls':
            logger.info(f"Detected .xls file (Excel 97-2003 format): {filepath}")
            logger.info(f"Converting to .xlsx format for processing...")
            try:
                self.filepath = self._convert_xls_to_xlsx(filepath)
                self.extension = '.xlsx'
                logger.info(f"✓ Conversion successful, processing as .xlsx")
            except ValueError as ve:
                # Conversion failed - provide helpful error
                logger.error(f"❌ .xls conversion failed: {str(ve)}")
                raise
            except Exception as e:
                logger.error(f"❌ Unexpected error converting .xls: {str(e)}")
                raise ValueError(f"Failed to process .xls file. Please save the file as .xlsx format in Excel/LibreOffice and upload again.")
        else:
            self.filepath = filepath
            self.extension = self.original_extension
        
        self.filename = os.path.basename(self.filepath)
        self.workbook = None
        self.images_cache = {}  # Cache for extracted images
        
    def _convert_xls_to_xlsx(self, xls_filepath):
        """
        Convert .xls file to .xlsx format for uniform processing
        
        Args:
            xls_filepath: Path to .xls file
            
        Returns:
            str: Path to converted .xlsx file
        """
        try:
            # Create output path (same directory, .xlsx extension)
            xlsx_filepath = xls_filepath.replace('.xls', '_converted.xlsx')
            
            # Try reading with xlrd
            try:
                logger.info(f"Attempting to read .xls file with xlrd: {xls_filepath}")
                xls_file = pd.ExcelFile(xls_filepath, engine='xlrd')
                logger.info(f"Successfully opened .xls file with {len(xls_file.sheet_names)} sheets")
            except Exception as xlrd_error:
                error_msg = str(xlrd_error).lower()
                logger.error(f"xlrd failed to read .xls file: {str(xlrd_error)}")
                
                # Provide specific error messages based on the error type
                if 'unsupported format' in error_msg or 'not a valid' in error_msg:
                    raise ValueError("The .xls file format is not valid or corrupted. Please open the file in Excel and save it as .xlsx format.")
                elif 'password' in error_msg or 'encrypted' in error_msg:
                    raise ValueError("The .xls file is password protected or encrypted. Please remove the password and save as .xlsx format.")
                else:
                    raise ValueError(f"Cannot read .xls file (error: {str(xlrd_error)}). Please save the file as .xlsx format in Excel and upload again.")
            
            # Create Excel writer for .xlsx
            logger.info(f"Converting sheets to .xlsx format...")
            with pd.ExcelWriter(xlsx_filepath, engine='openpyxl') as writer:
                for sheet_name in xls_file.sheet_names:
                    try:
                        logger.info(f"Converting sheet: {sheet_name}")
                        # Read each sheet
                        df = pd.read_excel(xls_file, sheet_name=sheet_name, header=None)
                        # Write to xlsx
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    except Exception as sheet_error:
                        logger.error(f"Error converting sheet {sheet_name}: {str(sheet_error)}")
                        # Continue with other sheets
                        continue
            
            # Verify the converted file exists
            if not os.path.exists(xlsx_filepath) or os.path.getsize(xlsx_filepath) == 0:
                raise ValueError("Conversion produced an empty or invalid file. Please save as .xlsx format manually.")
            
            logger.info(f"✓ Successfully converted .xls to .xlsx: {xlsx_filepath}")
            return xlsx_filepath
            
        except ValueError:
            # Re-raise ValueError with our custom message
            raise
        except Exception as e:
            logger.error(f"Unexpected error during .xls to .xlsx conversion: {str(e)}")
            raise ValueError(f"Cannot convert .xls file to .xlsx format. Please open the file in Excel and save it as .xlsx format, then upload again.")
    
    def _extract_images_from_sheet(self, sheet, output_dir):
        """
        Extract all images from an Excel sheet and map them to their cell positions
        Note: Only works with .xlsx files (openpyxl). .xls files don't support image extraction.
        Supports PNG, JPEG, GIF, BMP, WMF, EMF formats (WMF/EMF converted to PNG).
        
        Args:
            sheet: openpyxl worksheet object
            output_dir: Directory to save extracted images
            
        Returns:
            dict: Mapping of row numbers to image paths (row-based mapping)
        """
        images_dir = os.path.join(output_dir, 'imgs')
        os.makedirs(images_dir, exist_ok=True)
        
        row_images = {}  # Map row numbers to images
        conversion_attempts = []
        
        if not hasattr(sheet, '_images') or not sheet._images:
            logger.info(f"No images found in sheet '{sheet.title}'")
            return row_images
        
        total_images = len(sheet._images)
        logger.info(f"Found {total_images} images in sheet '{sheet.title}'")
        
        for idx, img in enumerate(sheet._images):
            try:
                # Check image format and detect WMF/EMF
                img_format = None
                is_wmf_emf = False
                img_data = None
                
                if hasattr(img, 'format'):
                    img_format = img.format
                
                # Get image data once and reuse it
                if hasattr(img, '_data'):
                    try:
                        img_data = img._data()  # Get data once
                        # Detect format from data
                        if img_data[:4] == b'\x89PNG':
                            img_format = 'png'
                        elif img_data[:2] == b'\xff\xd8':
                            img_format = 'jpeg'
                        elif img_data[:3] == b'GIF':
                            img_format = 'gif'
                        elif img_data[:2] == b'BM':
                            img_format = 'bmp'
                        elif img_data[:4] == b'\xd7\xcd\xc6\x9a' or img_data[:2] == b'\x01\x00':
                            img_format = 'wmf'
                            is_wmf_emf = True
                            logger.info(f"Image {idx} is WMF format - attempting conversion to PNG")
                        elif img_data[:4] == b'\x01\x00\x00\x00':
                            img_format = 'emf'
                            is_wmf_emf = True
                            logger.info(f"Image {idx} is EMF format - attempting conversion to PNG")
                    except Exception as data_error:
                        logger.debug(f"Could not read image {idx} data: {data_error}")
                
                # Get image anchor (position)
                anchor = None
                if hasattr(img, 'anchor'):
                    if hasattr(img.anchor, '_from'):
                        anchor = img.anchor._from
                    elif hasattr(img.anchor, 'from'):
                        anchor = img.anchor['from']
                
                if anchor:
                    # Get the row where image is anchored (1-indexed in Excel)
                    if hasattr(anchor, 'row'):
                        row_num = anchor.row + 1  # Convert to 1-indexed
                    elif hasattr(anchor, 'rowOff'):
                        row_num = anchor.rowOff + 1
                    else:
                        row_num = 0
                    
                    # Save image
                    img_filename = f"{sheet.title.replace(' ', '_')}_row{row_num}_img{idx}.png"
                    img_path = os.path.join(images_dir, img_filename)
                    
                    # Handle WMF/EMF conversion
                    if is_wmf_emf and img_data:
                        success = convert_wmf_emf_to_png(img_data, img_path)
                        if success:
                            conversion_attempts.append((idx, img_format.upper(), 'success'))
                            logger.info(f"✓ Converted {img_format.upper()} image {idx} at row {row_num}: {img_filename}")
                        else:
                            conversion_attempts.append((idx, img_format.upper(), 'failed'))
                            logger.warning(f"✗ Failed to convert {img_format.upper()} image {idx} - skipping")
                            continue
                    else:
                        # Save standard image formats (data already loaded)
                        try:
                            if img_data:
                                # Use the already-loaded data
                                with open(img_path, 'wb') as f:
                                    f.write(img_data)
                                logger.info(f"Extracted image {idx} at row {row_num}: {img_filename}")
                            elif hasattr(img, 'ref'):
                                # Handle embedded images
                                img_ref = img.ref
                                if hasattr(img_ref, '_data'):
                                    with open(img_path, 'wb') as f:
                                        f.write(img_ref())
                                    logger.info(f"Extracted image {idx} at row {row_num}: {img_filename}")
                            else:
                                logger.warning(f"Could not extract image {idx} - no data available")
                                continue
                        except Exception as save_error:
                            logger.error(f"Failed to save image {idx}: {save_error}")
                            continue
                    
                    # Store relative path mapped to row number
                    rel_path = f"imgs/{img_filename}"
                    
                    if row_num not in row_images:
                        row_images[row_num] = []
                    
                    # Only add if not already present (prevent duplicates)
                    if rel_path not in row_images[row_num]:
                        row_images[row_num].append(rel_path)
                    
                else:
                    logger.warning(f"Could not determine anchor for image {idx}")
                    
            except Exception as e:
                logger.error(f"Error extracting image {idx}: {e}", exc_info=True)
        
        # Log summary
        extracted_count = len([img for imgs in row_images.values() for img in imgs])
        logger.info(f"✓ Successfully extracted {extracted_count} images from {total_images} total")
        
        if conversion_attempts:
            successful = [c for c in conversion_attempts if c[2] == 'success']
            failed = [c for c in conversion_attempts if c[2] == 'failed']
            
            if successful:
                logger.info(f"✓ Converted {len(successful)} WMF/EMF image(s) to PNG:")
                for idx, fmt, _ in successful:
                    logger.info(f"  - Image {idx}: {fmt} → PNG")
            
            if failed:
                logger.warning(f"⚠ Failed to convert {len(failed)} WMF/EMF image(s):")
                for idx, fmt, _ in failed:
                    logger.warning(f"  - Image {idx}: {fmt} conversion failed")
                logger.warning(f"  Note: WMF/EMF conversion requires PIL, ImageMagick, or LibreOffice")
        
        return row_images
    
    def _get_cell_value_with_wrapping(self, cell):
        """
        Get cell value with proper text wrapping for long content
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            str: Formatted cell value
        """
        if cell.value is None:
            return ''
        
        text = str(cell.value).strip()
        
        # Check if text is long and needs wrapping
        if len(text) > 50:
            # Wrap text at reasonable points (spaces, punctuation)
            words = text.split()
            lines = []
            current_line = []
            current_length = 0
            
            for word in words:
                if current_length + len(word) + 1 <= 60:  # Target line length
                    current_line.append(word)
                    current_length += len(word) + 1
                else:
                    if current_line:
                        lines.append(' '.join(current_line))
                    current_line = [word]
                    current_length = len(word)
            
            if current_line:
                lines.append(' '.join(current_line))
            
            return '<br>'.join(lines)
        
        return text
        
    def extract_all_sheets(self, output_dir=None, session_id=None, file_id=None):
        """
        Extract data from all sheets in Excel file with smart table detection and images
        
        Args:
            output_dir: Directory to save extracted images
            session_id: Session ID for image URL paths
            file_id: File ID for image URL paths
            
        Returns:
            dict: Dictionary with sheet names as keys and data as values
        """
        try:
            logger.info(f"Extracting data from Excel file: {self.filename}")
            
            # Load workbook for image extraction (.xlsx format)
            if self.workbook is None:
                try:
                    logger.info(f"Loading workbook with openpyxl: {self.filepath}")
                    
                    # Use openpyxl with patched named range handling
                    from openpyxl import load_workbook
                    from openpyxl.reader.workbook import WorkbookParser
                    import warnings
                    
                    # Monkey-patch assign_names BEFORE loading to skip invalid definitions
                    original_assign_names = WorkbookParser.assign_names
                    
                    def patched_assign_names(self):
                        """Assign names with error handling for invalid definitions"""
                        try:
                            original_assign_names(self)
                        except (ValueError, Exception) as e:
                            # Skip invalid named ranges (like #N/A in print titles)
                            logger.warning(f"Skipping invalid named range during workbook load: {str(e)}")
                    
                    WorkbookParser.assign_names = patched_assign_names
                    
                    # Suppress warnings about invalid print areas
                    with warnings.catch_warnings():
                        warnings.filterwarnings('ignore', message='Print area cannot be set')
                        
                        try:
                            self.workbook = load_workbook(self.filepath, data_only=False)
                            logger.info(f"✓ Workbook loaded successfully")
                        finally:
                            # Restore original method
                            WorkbookParser.assign_names = original_assign_names
                            
                except Exception as wb_error:
                    error_msg = str(wb_error)
                    logger.error(f"Failed to load workbook: {error_msg}")
                    
                    # Provide specific error messages based on the error type
                    if 'XML' in error_msg or 'xml' in error_msg:
                        raise ValueError("Cannot read Excel file: The file appears to be corrupted or has XML structure issues. Please try:\n1. Open the file in Excel\n2. Save As > Excel Workbook (.xlsx)\n3. Upload the newly saved file")
                    elif 'zip' in error_msg.lower() or 'corrupt' in error_msg.lower():
                        raise ValueError("Cannot read Excel file: The file appears to be corrupted or damaged. Please repair the file in Excel (File > Info > Check for Issues) or recreate it.")
                    elif 'password' in error_msg.lower() or 'encrypted' in error_msg.lower():
                        raise ValueError("Cannot read Excel file: The file is password protected. Please remove the password and upload again.")
                    else:
                        raise ValueError(f"Cannot read Excel file: {error_msg}. Please ensure the file is a valid .xlsx format.")
            
            # Read all sheets
            try:
                logger.info(f"Reading sheets with pandas...")
                # Use openpyxl with patched named range handling
                from openpyxl import load_workbook
                from openpyxl.reader.workbook import WorkbookParser
                import warnings
                
                # Monkey-patch assign_names BEFORE loading to skip invalid definitions
                original_assign_names = WorkbookParser.assign_names
                
                def patched_assign_names(self):
                    """Assign names with error handling for invalid definitions"""
                    try:
                        original_assign_names(self)
                    except (ValueError, Exception) as e:
                        # Skip invalid named ranges (like #N/A in print titles)
                        logger.warning(f"Skipping invalid named range during sheet reading: {str(e)}")
                
                WorkbookParser.assign_names = patched_assign_names
                
                # Suppress warnings about invalid print areas
                with warnings.catch_warnings():
                    warnings.filterwarnings('ignore', message='Print area cannot be set')
                    
                    try:
                        wb = load_workbook(self.filepath, data_only=True, keep_links=False)
                        sheet_names = wb.sheetnames
                        wb.close()
                        logger.info(f"✓ Found {len(sheet_names)} sheets: {sheet_names}")
                    finally:
                        # Restore original method
                        WorkbookParser.assign_names = original_assign_names
                        
            except Exception as pd_error:
                error_msg = str(pd_error)
                logger.error(f"Failed to read file: {error_msg}")
                raise ValueError(f"Cannot read Excel file structure: {error_msg}. Please save the file as a new .xlsx file in Excel and try again.")
            
            results = {}
            for sheet_name in sheet_names:
                try:
                    sheet_data = self.extract_sheet(sheet_name, output_dir=output_dir, session_id=session_id, file_id=file_id)
                    
                    # Only include sheets that have data
                    if not sheet_data['empty']:
                        results[sheet_name] = sheet_data
                        logger.info(f"Sheet '{sheet_name}': {sheet_data['shape'][0]} rows, {sheet_data['shape'][1]} columns, {sheet_data.get('image_count', 0)} images")
                    else:
                        logger.info(f"Sheet '{sheet_name}': Skipped (empty or no valid data)")
                    
                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}': {e}")
                    results[sheet_name] = {
                        'error': str(e),
                        'data': [],
                        'html': f'<p>Error reading sheet: {str(e)}</p>',
                        'markdown': f'Error reading sheet: {str(e)}',
                        'images': {}
                    }
            
            return results
            
        except Exception as e:
            logger.error(f"Error reading Excel file {self.filename}: {e}")
            raise
    
    def _detect_table_start(self, df):
        """
        Detect where the actual table starts by finding the header row
        Looks for rows with multiple non-empty values that look like headers
        """
        # Common header keywords for BOQ/offer tables
        header_keywords = [
            'sn', 's.n', 'serial', 'item', 'description', 'desc', 
            'quantity', 'qty', 'unit', 'rate', 'price', 'amount', 
            'total', 'location', 'image', 'indicative', 'material'
        ]
        
        for idx, row in df.iterrows():
            # Count non-null values
            non_null = row.notna().sum()
            
            # Check if this row looks like a header
            if non_null >= 3:  # At least 3 columns should have values
                row_str = ' '.join(str(val).lower() for val in row if pd.notna(val))
                
                # Check if any header keywords are present
                keyword_count = sum(1 for keyword in header_keywords if keyword in row_str)
                
                if keyword_count >= 2:  # At least 2 header keywords found
                    logger.info(f"Detected table header at row {idx}: {row.tolist()}")
                    return idx
        
        # If no clear header found, return 0
        return 0
    
    def _is_product_table(self, df):
        """
        Validate if the dataframe contains a proper product-centric table structure.
        Returns True only if it has the essential columns for a product table.
        """
        if df.empty or len(df.columns) < 3:
            return False
        
        # Convert all column names to lowercase for comparison
        columns_lower = [str(col).lower().strip() for col in df.columns]
        columns_str = ' '.join(columns_lower)
        
        # Essential columns that MUST be present in a product table
        essential_keywords = {
            'item_id': ['sn', 's.n', 'sl.no', 'serial', 'item', 's no', 's.no', 'no.', 'no'],
            'description': ['description', 'desc', 'discription', 'item description', 'product', 'material', 'particulars', 'specification'],
            'pricing': ['rate', 'price', 'unit rate', 'unit price', 'amount', 'total', 'value', 'cost']
        }
        
        # Check for essential columns
        has_item_id = any(keyword in columns_str for keyword in essential_keywords['item_id'])
        has_description = any(keyword in columns_str for keyword in essential_keywords['description'])
        has_pricing = any(keyword in columns_str for keyword in essential_keywords['pricing'])
        
        # A valid product table MUST have at least: item identifier + description + pricing
        if not (has_item_id and has_description and has_pricing):
            logger.info(f"Sheet rejected: Missing essential product columns. Has ID:{has_item_id}, Desc:{has_description}, Price:{has_pricing}")
            return False
        
        # Additional validation: check if the data rows look like product data
        # Skip if it's mostly non-product data (project info, summaries, etc.)
        non_product_keywords = [
            'project', 'client', 'supplier', 'date', 'particulars', 
            'conversion', 'country', 'freight', 'insurance', 'customs',
            'clearance', 'add:', 'less:', 'total containers', 'material cost'
        ]
        
        # Check first few non-header rows for product-like content
        sample_rows = df.head(10)
        non_product_count = 0
        
        for _, row in sample_rows.iterrows():
            row_str = ' '.join(str(val).lower() for val in row if pd.notna(val))
            if any(keyword in row_str for keyword in non_product_keywords):
                non_product_count += 1
        
        # If more than 60% of sample rows contain non-product keywords, reject
        if non_product_count > len(sample_rows) * 0.6:
            logger.info(f"Sheet rejected: Contains too much non-product data ({non_product_count}/{len(sample_rows)} rows)")
            return False
        
        logger.info("Sheet validated as product table")
        return True
    
    def _clean_dataframe(self, df):
        """
        Clean dataframe by removing empty rows and columns while preserving _excel_row
        """
        # Store _excel_row if it exists
        excel_rows = df['_excel_row'].copy() if '_excel_row' in df.columns else None
        
        # Store column names before cleaning
        original_columns = df.columns.tolist()
        
        # Remove _excel_row temporarily for cleaning
        if '_excel_row' in df.columns:
            df = df.drop(columns=['_excel_row'])
            original_columns.remove('_excel_row')
        
        # Remove completely empty rows
        empty_mask = df.isna().all(axis=1)
        df = df[~empty_mask]
        
        # Remove completely empty columns, BUT preserve image columns
        # Image columns often appear empty because pandas can't read embedded images
        image_keywords = ['image', 'img', 'picture', 'photo', 'indicative']
        columns_to_keep = []
        
        for col in df.columns:
            col_lower = str(col).lower()
            # Keep if it's an image column or has any non-null values
            if any(keyword in col_lower for keyword in image_keywords) or df[col].notna().any():
                columns_to_keep.append(col)
        
        df = df[columns_to_keep]
        
        # Restore _excel_row if it existed
        if excel_rows is not None:
            df['_excel_row'] = excel_rows[~empty_mask].values
        
        return df
    
    def _is_valid_table_row(self, row, headers):
        """
        Check if a row is a valid data row (not a header repetition or empty)
        """
        # Skip _excel_row column if present
        actual_values = [val for val, col in zip(row, headers) if col != '_excel_row']
        
        # Convert row to strings for comparison
        row_str = [str(val).lower().strip() for val in actual_values]
        header_str = [str(h).lower().strip() for h in headers if h != '_excel_row']
        
        # Check if row is a header repetition
        if row_str == header_str:
            return False
        
        # Check if row has at least one non-empty value (very permissive)
        non_empty = sum(1 for val in actual_values if pd.notna(val) and str(val).strip() not in ['', 'nan', 'none'])
        
        return non_empty >= 1  # At least 1 column should have a value
    
    def extract_sheet(self, sheet_name=0, output_dir=None, session_id=None, file_id=None):
        """
        Extract data from specific sheet with smart table detection and image extraction
        
        Args:
            sheet_name: Sheet name (str) or index (int). Default is first sheet (0)
            output_dir: Directory to save extracted images
            session_id: Session ID for image URL paths
            file_id: File ID for image URL paths
            
        Returns:
            dict: Extracted data in multiple formats
        """
        try:
            # Load workbook with openpyxl for image extraction
            if self.workbook is None:
                self.workbook = openpyxl.load_workbook(self.filepath, data_only=False)
            
            # Get the worksheet
            if isinstance(sheet_name, int):
                ws = self.workbook.worksheets[sheet_name]
                actual_sheet_name = ws.title
            else:
                ws = self.workbook[sheet_name]
                actual_sheet_name = sheet_name
            
            # Check for defined print area to limit data extraction
            print_area = None
            try:
                if ws.print_area:
                    print_area = ws.print_area
                    logger.info(f"Found print area for sheet '{actual_sheet_name}': {print_area}")
            except Exception as e:
                logger.debug(f"No print area defined for sheet '{actual_sheet_name}': {e}")
            
            # Extract images if output_dir provided
            cell_images = {}
            if output_dir:
                # First, extract standard images via openpyxl
                cell_images = self._extract_images_from_sheet(ws, output_dir)
                
                # Then, extract WMF/EMF images directly from ZIP (bypasses openpyxl filtering)
                wmf_images = extract_wmf_images_from_excel_zip(
                    self.filepath, 
                    os.path.join(output_dir, 'imgs'),
                    sheet_name=actual_sheet_name
                )
                
                # Merge WMF images into cell_images mapping
                if wmf_images:
                    logger.info(f"✓ Extracted {len(wmf_images)} rows with WMF/EMF images")
                    for row_num, img_paths in wmf_images.items():
                        if row_num in cell_images:
                            cell_images[row_num].extend(img_paths)
                        else:
                            cell_images[row_num] = img_paths
                
                # Update image paths to include session/file_id for web access
                if session_id and file_id and cell_images:
                    updated_images = {}
                    for row_num, img_paths in cell_images.items():
                        updated_images[row_num] = [
                            f"/outputs/{session_id}/{file_id}/{path}" for path in img_paths
                        ]
                    cell_images = updated_images
            
            # Read with header=None to get raw data first
            # Apply monkey-patch for invalid named ranges
            from openpyxl.reader.workbook import WorkbookParser
            import warnings
            
            # Monkey-patch assign_names BEFORE reading
            original_assign_names = WorkbookParser.assign_names
            
            def patched_assign_names(self):
                """Assign names with error handling for invalid definitions"""
                try:
                    original_assign_names(self)
                except (ValueError, Exception) as e:
                    # Skip invalid named ranges (like #N/A in print titles)
                    logger.warning(f"Skipping invalid named range during pandas read: {str(e)}")
            
            WorkbookParser.assign_names = patched_assign_names
            
            # Suppress warnings about invalid print areas
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Print area cannot be set')
                
                try:
                    # Determine the range to read based on print area
                    use_cols = None
                    skip_rows = 0
                    n_rows = None
                    
                    if print_area:
                        # Parse print area (e.g., "$A$1:$Z$100" or "Sheet1!$A$1:$Z$100")
                        try:
                            # Remove sheet name if present
                            area = print_area.split('!')[-1] if '!' in print_area else print_area
                            # Remove $ signs
                            area = area.replace('$', '')
                            # If it's a valid range (contains ':'), parse it
                            if ':' in area:
                                start_cell, end_cell = area.split(':')
                                
                                # Extract column letters and row numbers
                                import re
                                start_match = re.match(r'([A-Z]+)(\d+)', start_cell)
                                end_match = re.match(r'([A-Z]+)(\d+)', end_cell)
                                
                                if start_match and end_match:
                                    start_col = start_match.group(1)
                                    start_row = int(start_match.group(2))
                                    end_col = end_match.group(1)
                                    end_row = int(end_match.group(2))
                                    
                                    # Create column range (e.g., "A:G")
                                    use_cols = f"{start_col}:{end_col}"
                                    skip_rows = start_row - 1  # pandas uses 0-based indexing
                                    n_rows = end_row - start_row + 1
                                    
                                    logger.info(f"Using print area: cols={use_cols}, skip_rows={skip_rows}, nrows={n_rows}")
                        except Exception as e:
                            logger.warning(f"Could not parse print area '{print_area}': {e}")
                    
                    # Use xlrd engine for .xls files
                    if self.extension == '.xls':
                        df_raw = pd.read_excel(self.filepath, sheet_name=actual_sheet_name, header=None, engine='xlrd')
                    else:
                        # Read with or without range restriction
                        read_params = {
                            'sheet_name': actual_sheet_name,
                            'header': None,
                            'engine': 'openpyxl'
                        }
                        
                        if use_cols:
                            read_params['usecols'] = use_cols
                        if skip_rows > 0:
                            read_params['skiprows'] = skip_rows
                        if n_rows:
                            read_params['nrows'] = n_rows
                        
                        df_raw = pd.read_excel(self.filepath, **read_params)
                    
                    # Detect where the table starts
                    header_row = self._detect_table_start(df_raw)
                    
                    logger.info(f"Detected header at row {header_row}")
                    
                    # Re-read with proper header
                    # Use xlrd engine for .xls files
                    if self.extension == '.xls':
                        if header_row > 0:
                            df = pd.read_excel(self.filepath, sheet_name=actual_sheet_name, header=header_row, engine='xlrd')
                        else:
                            df = pd.read_excel(self.filepath, sheet_name=actual_sheet_name, engine='xlrd')
                    else:
                        # Re-read with same parameters but proper header
                        read_params['header'] = header_row if header_row > 0 else 0
                        df = pd.read_excel(self.filepath, **read_params)
                finally:
                    # Restore original method
                    WorkbookParser.assign_names = original_assign_names
            
            # Store the actual Excel row number for each dataframe row
            # This is critical for image mapping
            df['_excel_row'] = range(header_row + 2, header_row + 2 + len(df))
            
            # Clean the dataframe
            df = self._clean_dataframe(df)
            
            if df.empty:
                logger.warning(f"Sheet '{actual_sheet_name}' is empty after cleaning")
                return {
                    'data': [],
                    'html': '<p>No data found</p>',
                    'markdown': 'No data found',
                    'columns': [],
                    'shape': (0, 0),
                    'empty': True,
                    'sheet_name': actual_sheet_name,
                    'images': {}
                }
            
            # Validate if this is a proper product table
            if not self._is_product_table(df):
                logger.warning(f"Sheet '{actual_sheet_name}' does not contain a valid product table structure - skipping")
                return {
                    'data': [],
                    'html': '<p>Sheet does not contain product table data</p>',
                    'markdown': 'Sheet does not contain product table data',
                    'columns': [],
                    'shape': (0, 0),
                    'empty': True,
                    'sheet_name': actual_sheet_name,
                    'images': {},
                    'validation_message': 'Not a product table - missing essential columns (SN, Description, Price)'
                }
            
            # Get headers
            headers = df.columns.tolist()
            
            # Filter valid data rows
            valid_rows = []
            valid_excel_rows = []
            for idx, row in df.iterrows():
                if self._is_valid_table_row(row.values, headers):
                    # Store both the row data and the excel row number
                    row_without_excel = {k: v for k, v in row.items() if k != '_excel_row'}
                    valid_rows.append(row_without_excel)
                    # Keep track of excel row numbers
                    if '_excel_row' in row:
                        valid_excel_rows.append(row['_excel_row'])
                    else:
                        valid_excel_rows.append(idx + header_row + 2)
            
            if valid_rows:
                df = pd.DataFrame(valid_rows)
                # Restore _excel_row column
                df['_excel_row'] = valid_excel_rows
            
            # Reset index
            df = df.reset_index(drop=True)
            
            # Enhance data with images and preserve full text
            enhanced_data = []
            for idx, row in df.iterrows():
                row_dict = {}
                
                # Get the actual Excel row number
                excel_row = row.get('_excel_row', idx + header_row + 2)
                
                # Debug: Check if this row should have images
                if idx < 5:  # Log first 5 rows for debugging
                    logger.info(f"Row {idx}: excel_row={excel_row}, has_images={excel_row in cell_images}")
                
                for col_name, value in row.items():
                    # Skip internal columns
                    if col_name == '_excel_row':
                        continue
                    
                    # Check if this row has images
                    row_has_images = excel_row in cell_images
                    
                    # For columns that typically contain images (INDICATIVE IMAGE, IMAGE, etc.)
                    col_lower = str(col_name).lower()
                    is_image_column = any(keyword in col_lower for keyword in ['image', 'picture', 'photo', 'img'])
                    
                    if row_has_images and is_image_column:
                        # Add images from this row with click-to-enlarge functionality
                        # Deduplicate images by using a set
                        unique_images = list(dict.fromkeys(cell_images[excel_row]))  # Preserve order, remove duplicates
                        
                        img_html = ''.join([
                            f'<img src="{img}" class="table-thumbnail" '
                            f'style="max-width:80px; max-height:80px; cursor:pointer; margin:2px; object-fit:cover; border: 1px solid #ddd; border-radius: 4px;" '
                            f'onclick="openImageModal(this.src)" '
                            f'title="Click to enlarge" />'
                            for img in unique_images
                        ])
                        
                        # Combine with text if present
                        if pd.notna(value) and str(value).strip():
                            text_content = str(value).strip()
                            row_dict[col_name] = f"{img_html}<br>{text_content}"
                        else:
                            row_dict[col_name] = img_html
                    else:
                        # Preserve full text without aggressive wrapping
                        if pd.notna(value):
                            text = str(value).strip()
                            # Only wrap extremely long text (>200 chars)
                            if len(text) > 200:
                                row_dict[col_name] = self._wrap_text(text, max_length=120)
                            else:
                                row_dict[col_name] = text
                        else:
                            row_dict[col_name] = ''
                
                enhanced_data.append(row_dict)
            
            # Debug: Log sample of enhanced data
            if idx < 2:  # Log first 2 rows
                logger.info(f"Enhanced row {idx} data keys: {list(row_dict.keys())}")
                if 'INDICATIVE IMAGE' in row_dict:
                    img_val = str(row_dict['INDICATIVE IMAGE'])[:150]
                    logger.info(f"Enhanced row {idx} INDICATIVE IMAGE: {img_val}...")
            
            # Remove _excel_row from headers if present
            headers_clean = [h for h in headers if h != '_excel_row']
            
            # Create enhanced DataFrame
            df_enhanced = pd.DataFrame(enhanced_data)
            
            logger.info(f"Sheet '{actual_sheet_name}': Found {len(df_enhanced)} valid rows with {len(headers_clean)} columns")
            logger.info(f"Headers: {headers_clean}")
            logger.info(f"Extracted {len(cell_images)} images from {len(set(cell_images.keys()))} rows")
            
            # Debug: Log first and last SN values
            if not df_enhanced.empty and 'SN' in df_enhanced.columns:
                first_sn = df_enhanced['SN'].iloc[0] if len(df_enhanced) > 0 else 'N/A'
                last_sn = df_enhanced['SN'].iloc[-1] if len(df_enhanced) > 0 else 'N/A'
                logger.info(f"SN range: first={first_sn}, last={last_sn}, total_rows={len(df_enhanced)}")
            
            # Debug: Check if images are in enhanced data
            has_img_col = 'INDICATIVE IMAGE' in df_enhanced.columns
            if has_img_col and len(df_enhanced) > 0:
                first_img_val = str(df_enhanced['INDICATIVE IMAGE'].iloc[0])[:100]
                logger.info(f"First row INDICATIVE IMAGE value: {first_img_val}...")
            
            # Generate HTML with embedded images
            html = df_enhanced.to_html(index=False, classes='table table-striped', escape=False, na_rep='')
            
            return {
                'data': enhanced_data,
                'html': html,
                'markdown': df[headers_clean].to_markdown(index=False),  # Markdown without HTML
                'columns': headers_clean,
                'shape': (len(df), len(headers_clean)),
                'empty': df.empty,
                'sheet_name': actual_sheet_name,
                'images': cell_images,
                'image_count': len(cell_images)
            }
            
        except Exception as e:
            logger.error(f"Error extracting sheet '{sheet_name}' from {self.filename}: {e}")
            raise
    
    def _wrap_text(self, text, max_length=60):
        """
        Wrap long text for better display
        
        Args:
            text: Text to wrap
            max_length: Maximum characters per line
            
        Returns:
            str: Wrapped text with HTML breaks
        """
        if not text or len(str(text)) <= max_length:
            return str(text)
        
        text = str(text).strip()
        words = text.split()
        lines = []
        current_line = []
        current_length = 0
        
        for word in words:
            if current_length + len(word) + 1 <= max_length:
                current_line.append(word)
                current_length += len(word) + 1
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word)
        
        if current_line:
            lines.append(' '.join(current_line))
        
        return '<br>'.join(lines)
    
    def get_sheet_names(self):
        """
        Get list of all sheet names in Excel file
        
        Returns:
            list: Sheet names
        """
        try:
            # Use openpyxl with patched named range handling
            from openpyxl import load_workbook
            from openpyxl.reader.workbook import WorkbookParser
            import warnings
            
            # Monkey-patch assign_names BEFORE loading to skip invalid definitions
            original_assign_names = WorkbookParser.assign_names
            
            def patched_assign_names(self):
                """Assign names with error handling for invalid definitions"""
                try:
                    original_assign_names(self)
                except (ValueError, Exception) as e:
                    # Skip invalid named ranges (like #N/A in print titles)
                    logger.warning(f"Skipping invalid named range in get_sheet_names: {str(e)}")
            
            WorkbookParser.assign_names = patched_assign_names
            
            # Suppress warnings about invalid print areas
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Print area cannot be set')
                
                try:
                    wb = load_workbook(self.filepath, data_only=True, keep_links=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                    return sheet_names
                finally:
                    # Restore original method
                    WorkbookParser.assign_names = original_assign_names
                    
        except Exception as e:
            logger.error(f"Error getting sheet names from {self.filename}: {e}")
            raise
    
    def to_json(self, sheet_name=None):
        """
        Convert Excel data to JSON format
        
        Args:
            sheet_name: Specific sheet to convert, or None for all sheets
            
        Returns:
            str: JSON string
        """
        try:
            if sheet_name:
                data = self.extract_sheet(sheet_name)
                return json.dumps(data, indent=2, default=str)
            else:
                data = self.extract_all_sheets()
                return json.dumps(data, indent=2, default=str)
        except Exception as e:
            logger.error(f"Error converting to JSON: {e}")
            raise
    
    def validate_file(self):
        """
        Validate if file exists and is a valid Excel file
        
        Returns:
            tuple: (bool: is_valid, str: error_message)
        """
        if not os.path.exists(self.filepath):
            return False, "File not found"
        
        if self.extension not in ['.xlsx']:
            return False, f"Invalid file extension after conversion: {self.extension}"
        
        try:
            # Validate the (converted) xlsx file using openpyxl with patched named range handling
            from openpyxl import load_workbook
            from openpyxl.reader.workbook import WorkbookParser
            import warnings
            
            # Monkey-patch assign_names BEFORE loading to skip invalid definitions
            original_assign_names = WorkbookParser.assign_names
            
            def patched_assign_names(self):
                """Assign names with error handling for invalid definitions"""
                try:
                    original_assign_names(self)
                except (ValueError, Exception) as e:
                    # Skip invalid named ranges (like #N/A in print titles)
                    logger.warning(f"Skipping invalid named range: {str(e)}")
            
            WorkbookParser.assign_names = patched_assign_names
            
            # Suppress warnings about invalid print areas
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Print area cannot be set')
                
                try:
                    wb = load_workbook(self.filepath, data_only=True, keep_links=False)
                    sheet_count = len(wb.sheetnames)
                    wb.close()
                    logger.info(f"Successfully validated Excel file with {sheet_count} sheets")
                    return True, "Valid Excel file"
                finally:
                    # Restore original method
                    WorkbookParser.assign_names = original_assign_names
                    
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"Excel validation failed: {str(e)}")
            logger.error(f"Full traceback: {error_details}")
            return False, f"Cannot read Excel file: {str(e)}"
    
    def get_file_info(self):
        """
        Get basic information about Excel file
        
        Returns:
            dict: File information
        """
        try:
            # Use openpyxl with patched named range handling
            from openpyxl import load_workbook
            from openpyxl.reader.workbook import WorkbookParser
            import warnings
            
            # Monkey-patch assign_names BEFORE loading to skip invalid definitions
            original_assign_names = WorkbookParser.assign_names
            
            def patched_assign_names(self):
                """Assign names with error handling for invalid definitions"""
                try:
                    original_assign_names(self)
                except (ValueError, Exception) as e:
                    # Skip invalid named ranges (like #N/A in print titles)
                    logger.warning(f"Skipping invalid named range in get_file_info: {str(e)}")
            
            WorkbookParser.assign_names = patched_assign_names
            
            # Suppress warnings about invalid print areas
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Print area cannot be set')
                
                try:
                    wb = load_workbook(self.filepath, data_only=True, keep_links=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                finally:
                    # Restore original method
                    WorkbookParser.assign_names = original_assign_names
            
            file_size = os.path.getsize(self.filepath)
            
            info = {
                'filename': self.filename,
                'filepath': self.filepath,
                'extension': self.extension,
                'size_bytes': file_size,
                'size_mb': round(file_size / (1024 * 1024), 2),
                'sheet_count': len(sheet_names),
                'sheet_names': sheet_names
            }
            
            return info
            
        except Exception as e:
            logger.error(f"Error getting file info: {e}")
            raise


def process_excel_file(filepath, output_dir=None, session_id=None, file_id=None):
    """
    Convenience function to process an Excel file
    
    Args:
        filepath: Path to Excel file
        output_dir: Directory to save extracted images (optional)
        session_id: Session ID for image URL paths (optional)
        file_id: File ID for image URL paths (optional)
        
    Returns:
        dict: Complete extraction results
    """
    processor = ExcelProcessor(filepath)
    
    # Validate file first
    is_valid, message = processor.validate_file()
    if not is_valid:
        return {
            'success': False,
            'error': message,
            'filepath': filepath
        }
    
    try:
        # Get file info
        file_info = processor.get_file_info()
        
        # Extract all sheets with images
        sheets_data = processor.extract_all_sheets(output_dir=output_dir, session_id=session_id, file_id=file_id)
        
        # Count total images across all sheets
        total_images = sum(sheet.get('image_count', 0) for sheet in sheets_data.values())
        
        return {
            'success': True,
            'file_info': file_info,
            'sheets': sheets_data,
            'sheet_count': len(sheets_data),
            'image_count': total_images,
            'message': f"Successfully extracted {len(sheets_data)} sheet(s) with {total_images} image(s)"
        }
        
    except ValueError as ve:
        # User-friendly validation errors - pass them through as-is
        logger.error(f"Validation error processing Excel file: {ve}")
        return {
            'success': False,
            'error': str(ve),
            'filepath': filepath
        }
    except Exception as e:
        # Unexpected errors - provide generic message with details
        logger.error(f"Unexpected error processing Excel file: {e}")
        error_msg = str(e)
        if 'XML' in error_msg or 'xml' in error_msg:
            user_message = "Cannot read Excel file: The file has XML structure issues. Please open in Excel, save as a new .xlsx file, and upload again."
        elif 'corrupt' in error_msg.lower() or 'damaged' in error_msg.lower():
            user_message = "Cannot read Excel file: The file appears corrupted. Please repair or recreate the file in Excel."
        else:
            user_message = f"Cannot read Excel file: {error_msg}"
        
        return {
            'success': False,
            'error': user_message,
            'filepath': filepath
        }
